#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate the house-readiness planning workbook.

Source data:
  finance/planning/goals.md   — debt balances, payoff sequence, targets
  finance/planning/cash-flow.md — paycheck structure, saving power
  finance/profile.md           — pay cadence

Output:
  finance/planning/house_project.xlsx

Re-run after editing the ASSUMPTIONS section below to regenerate scenarios.
"""

import math
import os
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT = os.path.join(SCRIPT_DIR, "..", "planning", "house_project.xlsx")

# ═══════════════════════════════════════════════════════════════════════════
# ASSUMPTIONS  — edit these, then re-run to regenerate the workbook
# ═══════════════════════════════════════════════════════════════════════════

MODEL_START = date(2026, 7, 1)
PAYCHECK_ANCHOR = date(2026, 6, 26)        # next expected paycheck Friday

# ── Brennan ──
BRENNAN_SAVE_PER_PC   = 404.49             # half of $808.98 monthly goal-routing
BRENNAN_3PC_EXTRA     = 967.49             # extra main-checking deposit in 3-pc months
BRENNAN_RAISE         = 0.03               # annual raise %
BRENNAN_PROMO_DATE    = date(2028, 7, 1)   # estimated promotion
BRENNAN_PROMO_UPLIFT  = 0.10               # one-time promotion bump

# ── Bailey ──
BAILEY_CAR_BAL        = 5_773.20
BAILEY_CAR_RATE       = 0.0774
BAILEY_CAR_PMT        = 150.00             # monthly car payment
BAILEY_POST_CAR_SAVE  = 150.00             # contribution to joint goals after car payoff

# ── Current debts (as of 2026-06-23) ──
K401_BAL              = 5_007.10
K401_RATE             = 0.08
K401_DEDUCTION        = 208.40             # per paycheck, payroll deduction

CAR_BAL               = 22_106.32
CAR_RATE              = 0.0569
CAR_PMT               = 423.54             # monthly

STU_BAL               = 13_868.53
STU_RATE              = 0.0375             # blended
STU_PMT               = 300.00             # monthly

# ── Ring financing (0% APR) ──
RING_AMOUNT           = 2_500.00
RING_MONTHS           = 10
RING_DELAY            = 1                  # months after 401(k) payoff before purchase

# ── Savings targets ──
EFUND_TARGET          = 15_000.00
HOUSE_TARGET          = 60_000.00
COMBINED_TARGET       = 75_000.00          # e-fund + house in same HYSA
FORECAST_END_HYSA     = 100_000.00         # continue forecast until HYSA reaches this

# ── HYSA (Capital One savings — holds both e-fund and house fund) ──
HYSA_RATE             = 0.03               # APY, compounded monthly
HYSA_STARTING         = 511.71             # balance as of 2026-05-31

# ── Mortgage scenario ──
HOUSE_PRICE           = 400_000.00
DOWN_PCT              = 0.10
MORT_RATE             = 0.065
MORT_YEARS            = 30
PROP_TAX_RATE         = 0.012
INSURANCE_ANNUAL      = 2_400.00
PMI_RATE              = 0.005              # annual, on loan balance; drops at 20% equity
MAINT_MONTHLY         = 200.00
CURRENT_RENT          = 2_270.00

# ── Scenario presets ──
SCENARIOS = {
    "Conservative": {"raise": 0.02, "promo_uplift": 0.07, "mort_rate": 0.070,
                     "save_capture": 0.75, "hysa_rate": 0.0275},
    "Base":         {"raise": 0.03, "promo_uplift": 0.10, "mort_rate": 0.065,
                     "save_capture": 0.85, "hysa_rate": 0.03},
    "Optimistic":   {"raise": 0.04, "promo_uplift": 0.12, "mort_rate": 0.060,
                     "save_capture": 0.95, "hysa_rate": 0.0325},
}

# ═══════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════

def _pc_in_month(year, month):
    """Count bi-weekly paychecks falling in a calendar month."""
    first = date(year, month, 1)
    nxt = date(year + (1 if month == 12 else 0), (month % 12) + 1, 1)
    d = PAYCHECK_ANCHOR
    while d >= first:
        d -= timedelta(days=14)
    while d < first:
        d += timedelta(days=14)
    n = 0
    while d < nxt:
        n += 1
        d += timedelta(days=14)
    return n


def _raise_mult(md, annual_raise, promo_uplift):
    yrs = max(0, (md - MODEL_START).days / 365.25)
    m = (1 + annual_raise) ** yrs
    if md >= BRENNAN_PROMO_DATE:
        m *= (1 + promo_uplift)
    return m


def _mortgage_pmt(principal, annual_rate, years):
    r = annual_rate / 12
    n = years * 12
    if r == 0:
        return principal / n
    return principal * r * (1 + r)**n / ((1 + r)**n - 1)


# ═══════════════════════════════════════════════════════════════════════════
# FORECAST ENGINE
# ═══════════════════════════════════════════════════════════════════════════

def run_forecast(scenario_name="Base"):
    """Run the month-by-month forecast for a given scenario preset."""
    sc = SCENARIOS[scenario_name]
    ann_raise      = sc["raise"]
    promo_uplift   = sc["promo_uplift"]
    mort_rate      = sc["mort_rate"]
    save_capture   = sc["save_capture"]
    hysa_rate      = sc["hysa_rate"]

    rows = []

    k401_bal  = K401_BAL
    k401_done = False;  k401_dt = None
    k401_countdown = -1

    ring_bal  = 0.0;  ring_pmt = RING_AMOUNT / RING_MONTHS
    ring_on   = False; ring_done = False

    hysa      = HYSA_STARTING
    efund_done = False; efund_dt = None
    car_bal   = CAR_BAL; car_done = False; car_dt = None
    stu_bal   = STU_BAL; stu_done = False; stu_dt = None
    bailey_car_bal = BAILEY_CAR_BAL; bailey_car_done = False; bailey_car_dt = None
    debt_free_dt = None
    house_dt  = None
    hysa_100k_dt = None

    for i in range(120):                   # up to 10 years
        y = MODEL_START.year + (MODEL_START.month - 1 + i) // 12
        m = (MODEL_START.month - 1 + i) % 12 + 1
        md = date(y, m, 1)

        if hysa_100k_dt:
            break

        pc = _pc_in_month(y, m)
        rmult = _raise_mult(md, ann_raise, promo_uplift)

        # ── HYSA interest (compounds monthly on full balance) ──
        hysa_interest = hysa * hysa_rate / 12
        hysa += hysa_interest

        # ── Brennan saving power ──
        b_save = BRENNAN_SAVE_PER_PC * rmult * pc * save_capture
        if pc >= 3:
            b_save += BRENNAN_3PC_EXTRA * rmult * save_capture
        if k401_done:
            b_save += K401_DEDUCTION * pc          # payroll release (fixed $)
        if car_done:
            b_save += CAR_PMT                       # monthly release
        if stu_done:
            b_save += STU_PMT                       # monthly release

        # ── Bailey ──
        bl_save = 0.0
        if bailey_car_bal > 0:
            bci = bailey_car_bal * BAILEY_CAR_RATE / 12
            bailey_car_bal = max(0, bailey_car_bal - max(0, BAILEY_CAR_PMT - bci))
            if bailey_car_bal == 0 and not bailey_car_done:
                bailey_car_done = True; bailey_car_dt = md
        if bailey_car_done:
            bl_save = BAILEY_POST_CAR_SAVE

        total = b_save + bl_save
        cash  = total

        # ── Regular debt amortization (always runs) ──
        if car_bal > 0:
            ci = car_bal * CAR_RATE / 12
            car_bal = max(0, car_bal - max(0, CAR_PMT - ci))
            if car_bal == 0 and not car_done:
                car_done = True; car_dt = md

        if stu_bal > 0:
            si = stu_bal * STU_RATE / 12
            stu_bal = max(0, stu_bal - max(0, STU_PMT - si))
            if stu_bal == 0 and not stu_done:
                stu_done = True; stu_dt = md

        # ── Phase 1: 401(k) loan payoff ──
        phase = ""
        if not k401_done:
            phase = "401(k) Loan Payoff"
            ki = k401_bal * K401_RATE / 12
            k401_bal = k401_bal + ki - K401_DEDUCTION * pc
            hysa += cash; cash = 0
            if k401_bal <= 0:
                k401_bal = 0; k401_done = True; k401_dt = md
            elif hysa >= k401_bal:
                hysa -= k401_bal
                k401_bal = 0; k401_done = True; k401_dt = md

        # ── Phase 2/3: Ring + E-fund (HYSA toward $15k gate) ──
        if k401_done and not efund_done:
            if k401_countdown < 0:
                k401_countdown = 0
            k401_countdown += 1
            if not ring_on and not ring_done and k401_countdown > RING_DELAY:
                ring_on = True; ring_bal = RING_AMOUNT
            if ring_on and ring_bal > 0:
                phase = "Ring + E-Fund"
                rp = min(ring_bal, ring_pmt)
                ring_bal -= rp; cash = max(0, cash - rp)
                if ring_bal <= 0:
                    ring_bal = 0; ring_done = True; ring_on = False
            else:
                phase = "Emergency Fund"
            hysa += cash; cash = 0
            if hysa >= EFUND_TARGET:
                cash = hysa - EFUND_TARGET
                hysa = EFUND_TARGET; efund_done = True; efund_dt = md

        # ── Phase 4: Car extra payments ──
        if efund_done and car_bal > 0 and not car_done:
            phase = "Car Extra Payments"
            extra = min(car_bal, cash); car_bal -= extra; cash -= extra
            if car_bal <= 0:
                car_bal = 0; car_done = True; car_dt = md

        # ── Phase 5: Student extra payments ──
        if efund_done and car_done and stu_bal > 0 and not stu_done:
            phase = "Student Loan Extra"
            extra = min(stu_bal, cash); stu_bal -= extra; cash -= extra
            if stu_bal <= 0:
                stu_bal = 0; stu_done = True; stu_dt = md

        # Debt-free gate
        if k401_done and car_done and stu_done and not debt_free_dt:
            debt_free_dt = md

        # ── Phase 6: House fund (HYSA toward $75k combined target) ──
        if k401_done and car_done and stu_done and efund_done:
            phase = "House Fund"
            hysa += cash; cash = 0
            if hysa >= COMBINED_TARGET and not house_dt:
                house_dt = md
            if hysa >= FORECAST_END_HYSA and not hysa_100k_dt:
                hysa_100k_dt = md

        rows.append({
            'date': md, 'pc': pc,
            'b_save': b_save, 'bl_save': bl_save, 'total': total,
            'phase': phase, 'hysa_interest': hysa_interest,
            'k401': max(0, k401_bal),
            'ring': ring_bal, 'hysa': hysa,
            'car': car_bal, 'stu': stu_bal,
        })

    milestones = {
        '401(k) Loan Paid Off':       k401_dt,
        'Emergency Fund Complete':     efund_dt,
        'Car Loan Paid Off':           car_dt,
        'Student Loans Paid Off':      stu_dt,
        'Bailey Car Paid Off':         bailey_car_dt,
        'Fully Debt-Free':             debt_free_dt,
        'HYSA Reaches $75k':           house_dt,
        'HYSA Reaches $100k':          hysa_100k_dt,
    }

    # ── Mortgage affordability at house-readiness ──
    loan_amt = HOUSE_PRICE * (1 - DOWN_PCT)
    mo_pi = _mortgage_pmt(loan_amt, mort_rate, MORT_YEARS)
    mo_tax = HOUSE_PRICE * PROP_TAX_RATE / 12
    mo_ins = INSURANCE_ANNUAL / 12
    mo_pmi = loan_amt * PMI_RATE / 12 if DOWN_PCT < 0.20 else 0
    mo_maint = MAINT_MONTHLY
    total_housing = mo_pi + mo_tax + mo_ins + mo_pmi + mo_maint

    # Cumulative HYSA interest earned
    total_hysa_interest = sum(r['hysa_interest'] for r in rows)

    # Post-debt monthly housing budget (at house-ready date, if available)
    # This is the TOTAL monthly income available for housing: current rent
    # (redirected to mortgage) + freed debt payments + saving power with raises
    # + Bailey's contribution. The user is currently paying rent AND debts AND
    # routing $809/mo to goals simultaneously, so all of that is budget.
    if house_dt:
        rmult_h = _raise_mult(house_dt, ann_raise, promo_uplift)
        saving_power_at_ready = BRENNAN_SAVE_PER_PC * rmult_h * 2 * save_capture
        k401_release = K401_DEDUCTION * 2
        debt_releases = k401_release + CAR_PMT + STU_PMT
        bailey_at_ready = BAILEY_POST_CAR_SAVE
        post_debt_budget = (CURRENT_RENT + saving_power_at_ready
                            + debt_releases + bailey_at_ready)
    else:
        saving_power_at_ready = None
        debt_releases = None
        bailey_at_ready = None
        post_debt_budget = None

    mortgage_info = {
        'loan': loan_amt, 'pi': mo_pi, 'tax': mo_tax, 'ins': mo_ins,
        'pmi': mo_pmi, 'maint': mo_maint, 'total': total_housing,
        'post_debt_budget': post_debt_budget,
        'saving_power_at_ready': saving_power_at_ready,
        'debt_releases': debt_releases,
        'bailey_at_ready': bailey_at_ready,
        'mort_rate': mort_rate,
        'hysa_rate': hysa_rate,
        'total_hysa_interest': total_hysa_interest,
    }

    return rows, milestones, mortgage_info


# ═══════════════════════════════════════════════════════════════════════════
# EXCEL BUILDER
# ═══════════════════════════════════════════════════════════════════════════

# Styles
HDR_FONT   = Font(name="Calibri", bold=True, size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14)
SEC_FONT   = Font(name="Calibri", bold=True, size=12, color="1F4E79")
EDIT_FILL  = PatternFill("solid", fgColor="DAEEF3")   # light blue = editable
MILE_FILL  = PatternFill("solid", fgColor="C6EFCE")   # light green = milestone
WARN_FILL  = PatternFill("solid", fgColor="FFC7CE")   # light red = warning
HDR_FILL   = PatternFill("solid", fgColor="4472C4")   # dark blue header
HDR_FONT_W = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CUR_FMT  = '$#,##0.00'
CUR0_FMT = '$#,##0'
PCT_FMT  = '0.0%'
DATE_FMT = 'MMM YYYY'


def _write_assumption(ws, row, label, value, fmt=None, editable=False):
    """Write a label-value pair on the Assumptions sheet."""
    ws.cell(row=row, column=1, value=label).font = Font(name="Calibri", size=11)
    c = ws.cell(row=row, column=2, value=value)
    c.font = Font(name="Calibri", size=11)
    if fmt:
        c.number_format = fmt
    if editable:
        c.fill = EDIT_FILL
    return row + 1


def _section_header(ws, row, title):
    ws.cell(row=row, column=1, value=title).font = SEC_FONT
    return row + 1


def build_assumptions_sheet(wb):
    ws = wb.active
    ws.title = "Assumptions"
    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 40

    r = 1
    ws.cell(row=r, column=1, value="House Readiness Planning Model").font = TITLE_FONT
    r += 1
    ws.cell(row=r, column=1,
            value="Blue cells are editable scenario inputs. Re-run generate_house_model.py after changes.").font = Font(
        name="Calibri", size=10, italic=True, color="4472C4")
    r += 2

    r = _section_header(ws, r, "BRENNAN — INCOME & SAVING")
    r = _write_assumption(ws, r, "Saving power per paycheck", BRENNAN_SAVE_PER_PC, CUR_FMT, True)
    r = _write_assumption(ws, r, "Extra in 3-paycheck months", BRENNAN_3PC_EXTRA, CUR_FMT, True)
    r = _write_assumption(ws, r, "Annual raise assumption", BRENNAN_RAISE, PCT_FMT, True)
    r = _write_assumption(ws, r, "Promotion date", BRENNAN_PROMO_DATE, DATE_FMT, True)
    r = _write_assumption(ws, r, "Promotion uplift", BRENNAN_PROMO_UPLIFT, PCT_FMT, True)
    ws.cell(row=r-5, column=3,
            value="Source: cash-flow.md — $808.98/mo ÷ 2 paychecks").font = Font(
        name="Calibri", size=9, italic=True, color="808080")
    r += 1

    r = _section_header(ws, r, "BAILEY")
    r = _write_assumption(ws, r, "Car loan balance", BAILEY_CAR_BAL, CUR_FMT)
    r = _write_assumption(ws, r, "Car loan rate", BAILEY_CAR_RATE, PCT_FMT)
    r = _write_assumption(ws, r, "Car loan monthly payment", BAILEY_CAR_PMT, CUR_FMT)
    r = _write_assumption(ws, r, "Post-car contribution to joint goals", BAILEY_POST_CAR_SAVE, CUR_FMT)
    r += 1

    r = _section_header(ws, r, "CURRENT DEBTS (as of 2026-06-23)")
    r = _write_assumption(ws, r, "401(k) loan balance", K401_BAL, CUR_FMT)
    r = _write_assumption(ws, r, "401(k) loan rate", K401_RATE, PCT_FMT)
    r = _write_assumption(ws, r, "401(k) payroll deduction per paycheck", K401_DEDUCTION, CUR_FMT)
    ws.cell(row=r-3, column=3,
            value="Source: goals.md — debt inventory").font = Font(
        name="Calibri", size=9, italic=True, color="808080")
    r += 1
    r = _write_assumption(ws, r, "Car loan balance", CAR_BAL, CUR_FMT)
    r = _write_assumption(ws, r, "Car loan rate", CAR_RATE, PCT_FMT)
    r = _write_assumption(ws, r, "Car loan monthly payment", CAR_PMT, CUR_FMT)
    r += 1
    r = _write_assumption(ws, r, "Student loan balance (blended)", STU_BAL, CUR_FMT)
    r = _write_assumption(ws, r, "Student loan blended rate", STU_RATE, PCT_FMT)
    r = _write_assumption(ws, r, "Student loan monthly payment", STU_PMT, CUR_FMT)
    r += 1
    r = _write_assumption(ws, r, "Total tracked debt", K401_BAL + CAR_BAL + STU_BAL, CUR_FMT)
    r += 1

    r = _section_header(ws, r, "RING FINANCING")
    r = _write_assumption(ws, r, "Ring purchase amount", RING_AMOUNT, CUR_FMT, True)
    r = _write_assumption(ws, r, "Payoff window (months)", RING_MONTHS, None, True)
    r = _write_assumption(ws, r, "Months after 401(k) payoff to start", RING_DELAY, None, True)
    r += 1

    r = _section_header(ws, r, "SAVINGS TARGETS")
    r = _write_assumption(ws, r, "Emergency fund target", EFUND_TARGET, CUR0_FMT, True)
    r = _write_assumption(ws, r, "House fund target (down + closing)", HOUSE_TARGET, CUR0_FMT, True)
    r = _write_assumption(ws, r, "Combined HYSA target", COMBINED_TARGET, CUR0_FMT)
    ws.cell(row=r-1, column=3,
            value="E-fund + house share one HYSA account").font = Font(
        name="Calibri", size=9, italic=True, color="808080")
    r += 1

    r = _section_header(ws, r, "HYSA (CAPITAL ONE SAVINGS)")
    r = _write_assumption(ws, r, "Starting balance", HYSA_STARTING, CUR_FMT)
    r = _write_assumption(ws, r, "APY (compounded monthly)", HYSA_RATE, PCT_FMT, True)
    ws.cell(row=r-1, column=3,
            value="Varies by scenario: 2.75% / 3.0% / 3.25%").font = Font(
        name="Calibri", size=9, italic=True, color="808080")
    r += 1

    r = _section_header(ws, r, "MORTGAGE SCENARIO")
    r = _write_assumption(ws, r, "House price", HOUSE_PRICE, CUR0_FMT, True)
    r = _write_assumption(ws, r, "Down payment %", DOWN_PCT, PCT_FMT, True)
    r = _write_assumption(ws, r, "Mortgage rate", MORT_RATE, PCT_FMT, True)
    r = _write_assumption(ws, r, "Mortgage term (years)", MORT_YEARS, None, True)
    r = _write_assumption(ws, r, "Property tax rate", PROP_TAX_RATE, PCT_FMT, True)
    r = _write_assumption(ws, r, "Homeowners insurance (annual)", INSURANCE_ANNUAL, CUR0_FMT, True)
    r = _write_assumption(ws, r, "PMI rate (annual, on loan)", PMI_RATE, PCT_FMT, True)
    r = _write_assumption(ws, r, "Monthly maintenance reserve", MAINT_MONTHLY, CUR0_FMT, True)
    r += 1
    r = _write_assumption(ws, r, "Current rent (for comparison)", CURRENT_RENT, CUR0_FMT)
    r += 1

    r = _section_header(ws, r, "PAYCHECK CALENDAR")
    r = _write_assumption(ws, r, "Paycheck anchor date", PAYCHECK_ANCHOR, "MMM DD, YYYY")
    r = _write_assumption(ws, r, "Frequency", "Bi-weekly (every 14 days)")
    r += 1

    r = _section_header(ws, r, "SCENARIO PRESETS")
    ws.cell(row=r, column=1, value="Preset").font = HDR_FONT
    ws.cell(row=r, column=2, value="Raise").font = HDR_FONT
    ws.cell(row=r, column=3, value="Promo Uplift").font = HDR_FONT
    ws.cell(row=r, column=4, value="Mort Rate").font = HDR_FONT
    ws.cell(row=r, column=5, value="Save Capture").font = HDR_FONT
    ws.cell(row=r, column=6, value="HYSA Rate").font = HDR_FONT
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    r += 1
    for name, sc in SCENARIOS.items():
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=sc["raise"]).number_format = PCT_FMT
        ws.cell(row=r, column=3, value=sc["promo_uplift"]).number_format = PCT_FMT
        ws.cell(row=r, column=4, value=sc["mort_rate"]).number_format = PCT_FMT
        ws.cell(row=r, column=5, value=sc["save_capture"]).number_format = PCT_FMT
        ws.cell(row=r, column=6, value=sc["hysa_rate"]).number_format = PCT_FMT
        r += 1

    return ws


def build_forecast_sheet(wb, rows, scenario_name):
    ws = wb.create_sheet(f"Forecast — {scenario_name}")

    headers = [
        ("Month", 12), ("PCs", 5), ("Brennan $", 13), ("Bailey $", 12),
        ("Total $", 13), ("Phase", 22),
        ("401(k) Loan", 13), ("Ring Bal", 11),
        ("HYSA Balance", 14), ("HYSA Interest", 13),
        ("Car Loan", 13), ("Student Loan", 13),
    ]
    for ci, (hdr, w) in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=hdr)
        c.font = HDR_FONT_W
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w

    num_cols = len(headers)
    for ri, row in enumerate(rows, 2):
        ws.cell(row=ri, column=1, value=row['date']).number_format = DATE_FMT
        ws.cell(row=ri, column=2, value=row['pc']).alignment = Alignment(horizontal="center")
        ws.cell(row=ri, column=3, value=row['b_save']).number_format = CUR_FMT
        ws.cell(row=ri, column=4, value=row['bl_save']).number_format = CUR_FMT
        ws.cell(row=ri, column=5, value=row['total']).number_format = CUR_FMT
        ws.cell(row=ri, column=6, value=row['phase'])
        ws.cell(row=ri, column=7, value=row['k401']).number_format = CUR_FMT
        ws.cell(row=ri, column=8, value=row['ring']).number_format = CUR_FMT
        ws.cell(row=ri, column=9, value=row['hysa']).number_format = CUR_FMT
        ws.cell(row=ri, column=10, value=row['hysa_interest']).number_format = CUR_FMT
        ws.cell(row=ri, column=11, value=row['car']).number_format = CUR_FMT
        ws.cell(row=ri, column=12, value=row['stu']).number_format = CUR_FMT

        # Highlight 3-paycheck months
        if row['pc'] >= 3:
            for col in range(1, num_cols + 1):
                ws.cell(row=ri, column=col).fill = PatternFill("solid", fgColor="FFF2CC")

        # Highlight milestone transitions (balance hits zero)
        for col_idx in (7, 8, 11, 12):
            val = ws.cell(row=ri, column=col_idx).value
            if ri > 2:
                prev = ws.cell(row=ri-1, column=col_idx).value
                if prev and prev > 0 and (val is None or val == 0):
                    ws.cell(row=ri, column=col_idx).fill = MILE_FILL

        # Highlight HYSA when it passes e-fund gate and combined target
        if row['hysa'] >= EFUND_TARGET:
            if ri == 2 or rows[ri-3]['hysa'] < EFUND_TARGET:
                ws.cell(row=ri, column=9).fill = MILE_FILL
        if row['hysa'] >= COMBINED_TARGET:
            if ri == 2 or rows[ri-3]['hysa'] < COMBINED_TARGET:
                ws.cell(row=ri, column=9).fill = MILE_FILL

        # Apply thin borders
        for col in range(1, num_cols + 1):
            ws.cell(row=ri, column=col).border = THIN_BORDER

    # Freeze header row
    ws.freeze_panes = "A2"
    return ws


def build_mortgage_sheet(wb, mortgage_info):
    ws = wb.create_sheet("Mortgage Affordability")
    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18

    r = 1
    ws.cell(row=r, column=1, value="Mortgage Affordability Analysis").font = TITLE_FONT
    r += 2

    r = _section_header(ws, r, "MONTHLY HOUSING COST BREAKDOWN")
    items = [
        ("House price", HOUSE_PRICE, CUR0_FMT),
        ("Down payment", HOUSE_PRICE * DOWN_PCT, CUR0_FMT),
        ("Loan amount", mortgage_info['loan'], CUR0_FMT),
        ("Mortgage rate", mortgage_info['mort_rate'], PCT_FMT),
        ("", None, None),
        ("Principal + Interest (P&I)", mortgage_info['pi'], CUR_FMT),
        ("Property tax (monthly)", mortgage_info['tax'], CUR_FMT),
        ("Homeowners insurance (monthly)", mortgage_info['ins'], CUR_FMT),
        ("PMI (monthly)", mortgage_info['pmi'], CUR_FMT),
        ("Maintenance reserve", mortgage_info['maint'], CUR_FMT),
        ("", None, None),
        ("TOTAL MONTHLY HOUSING COST", mortgage_info['total'], CUR_FMT),
    ]
    for label, val, fmt in items:
        if label == "":
            r += 1; continue
        c1 = ws.cell(row=r, column=1, value=label)
        if label.startswith("TOTAL"):
            c1.font = Font(name="Calibri", bold=True, size=12)
        c2 = ws.cell(row=r, column=2, value=val)
        if fmt:
            c2.number_format = fmt
        if label.startswith("TOTAL"):
            c2.font = Font(name="Calibri", bold=True, size=12)
            c2.fill = PatternFill("solid", fgColor="E2EFDA")
        r += 1
    r += 1

    r = _section_header(ws, r, "AFFORDABILITY COMPARISON")
    ws.cell(row=r, column=1, value="Current rent")
    ws.cell(row=r, column=2, value=CURRENT_RENT).number_format = CUR_FMT
    r += 1
    ws.cell(row=r, column=1, value="Projected housing cost")
    ws.cell(row=r, column=2, value=mortgage_info['total']).number_format = CUR_FMT
    r += 1
    delta = mortgage_info['total'] - CURRENT_RENT
    c1 = ws.cell(row=r, column=1, value="Difference vs. rent")
    c2 = ws.cell(row=r, column=2, value=delta)
    c2.number_format = CUR_FMT
    if delta > 0:
        c2.fill = WARN_FILL
        ws.cell(row=r, column=3,
                value=f"+${delta:,.0f}/mo more than rent").font = Font(
            name="Calibri", size=10, italic=True, color="C00000")
    else:
        c2.fill = MILE_FILL
        ws.cell(row=r, column=3,
                value=f"${abs(delta):,.0f}/mo less than rent").font = Font(
            name="Calibri", size=10, italic=True, color="006100")
    r += 2

    if mortgage_info['post_debt_budget']:
        r = _section_header(ws, r, "POST-DEBT HOUSING BUDGET")
        ws.cell(row=r, column=1,
                value="You currently pay rent AND debts AND save ~$809/mo simultaneously.").font = Font(
            name="Calibri", size=10, italic=True, color="808080")
        r += 1
        ws.cell(row=r, column=1,
                value="Post-debt, all of that income is available for housing:").font = Font(
            name="Calibri", size=10, italic=True, color="808080")
        r += 2

        budget = mortgage_info['post_debt_budget']
        budget_items = [
            ("Current rent (redirects to mortgage)", CURRENT_RENT),
            ("Freed debt payments (401k + car + student + Bailey car)",
             mortgage_info['debt_releases']),
            ("Saving power with raises (2-paycheck baseline)",
             mortgage_info['saving_power_at_ready']),
            ("Bailey contribution with raises",
             mortgage_info['bailey_at_ready']),
        ]
        for label, val in budget_items:
            ws.cell(row=r, column=1, value=label)
            ws.cell(row=r, column=2, value=val).number_format = CUR_FMT
            r += 1
        r += 1
        ws.cell(row=r, column=1, value="MAXIMUM MONTHLY HOUSING BUDGET").font = Font(
            name="Calibri", bold=True, size=12)
        c = ws.cell(row=r, column=2, value=budget)
        c.number_format = CUR_FMT
        c.font = Font(name="Calibri", bold=True, size=12)
        c.fill = PatternFill("solid", fgColor="E2EFDA")
        r += 2

        ws.cell(row=r, column=1, value="Housing cost as % of max budget")
        ratio = mortgage_info['total'] / budget if budget > 0 else 0
        c = ws.cell(row=r, column=2, value=ratio)
        c.number_format = PCT_FMT
        if ratio > 0.80:
            c.fill = WARN_FILL
        else:
            c.fill = MILE_FILL
        r += 1
        ws.cell(row=r, column=1, value="Remaining monthly after housing")
        remaining = budget - mortgage_info['total']
        c = ws.cell(row=r, column=2, value=remaining)
        c.number_format = CUR_FMT
        c.fill = MILE_FILL if remaining > 500 else WARN_FILL
        r += 1
        ws.cell(row=r, column=1,
                value="Note: uses 2-paycheck month baseline. 3-paycheck months add ~$967 extra.").font = Font(
            name="Calibri", size=9, italic=True, color="808080")
        r += 2

    # PMI drop-off note
    r = _section_header(ws, r, "PMI NOTE")
    if DOWN_PCT < 0.20:
        pmi_equity = HOUSE_PRICE * 0.20
        needed = pmi_equity - HOUSE_PRICE * DOWN_PCT
        ws.cell(row=r, column=1,
                value=f"PMI drops off at 20% equity (${pmi_equity:,.0f}). Need ${needed:,.0f} in equity beyond down payment.")
        r += 1
        ws.cell(row=r, column=1,
                value=f"PMI savings when removed: ${mortgage_info['pmi']:,.2f}/month")
    else:
        ws.cell(row=r, column=1, value="No PMI required — down payment ≥ 20%.")

    return ws


def build_summary_sheet(wb, all_results):
    ws = wb.create_sheet("Summary")
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 34
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16

    r = 1
    ws.cell(row=r, column=1, value="House Readiness Summary").font = TITLE_FONT
    r += 2

    # ── Milestone comparison across scenarios ──
    r = _section_header(ws, r, "MILESTONE DATES BY SCENARIO")

    milestone_names = [
        '401(k) Loan Paid Off', 'Emergency Fund Complete',
        'Car Loan Paid Off', 'Student Loans Paid Off',
        'Bailey Car Paid Off',
        'Fully Debt-Free', 'HYSA Reaches $75k', 'HYSA Reaches $100k',
    ]
    scenario_names = list(all_results.keys())

    ws.cell(row=r, column=1, value="Milestone").font = HDR_FONT_W
    ws.cell(row=r, column=1).fill = HDR_FILL
    ws.cell(row=r, column=1).border = THIN_BORDER
    for si, sname in enumerate(scenario_names, 2):
        c = ws.cell(row=r, column=si, value=sname)
        c.font = HDR_FONT_W; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = THIN_BORDER
    r += 1

    for mname in milestone_names:
        ws.cell(row=r, column=1, value=mname).border = THIN_BORDER
        for si, sname in enumerate(scenario_names, 2):
            _, milestones, _ = all_results[sname]
            dt = milestones.get(mname)
            c = ws.cell(row=r, column=si, value=dt if dt else "—")
            if dt:
                c.number_format = "MMM YYYY"
            c.alignment = Alignment(horizontal="center")
            c.border = THIN_BORDER
            if mname == 'HYSA Reaches $75k' and dt:
                c.fill = MILE_FILL
            if mname == 'HYSA Reaches $100k' and dt:
                c.fill = MILE_FILL
        r += 1
    r += 1

    # ── Monthly housing cost comparison ──
    r = _section_header(ws, r, "MONTHLY HOUSING COST BY SCENARIO")
    ws.cell(row=r, column=1, value="Component").font = HDR_FONT_W
    ws.cell(row=r, column=1).fill = HDR_FILL
    ws.cell(row=r, column=1).border = THIN_BORDER
    for si, sname in enumerate(scenario_names, 2):
        c = ws.cell(row=r, column=si, value=sname)
        c.font = HDR_FONT_W; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center"); c.border = THIN_BORDER
    r += 1

    components = ['pi', 'tax', 'ins', 'pmi', 'maint', 'total', 'hysa_rate', 'total_hysa_interest']
    comp_labels = {
        'pi': 'Principal + Interest', 'tax': 'Property Tax',
        'ins': 'Insurance', 'pmi': 'PMI', 'maint': 'Maintenance',
        'total': 'TOTAL HOUSING',
        'hysa_rate': 'HYSA APY', 'total_hysa_interest': 'Total HYSA Interest Earned',
    }
    for comp in components:
        ws.cell(row=r, column=1, value=comp_labels[comp]).border = THIN_BORDER
        if comp == 'total':
            ws.cell(row=r, column=1).font = HDR_FONT
        for si, sname in enumerate(scenario_names, 2):
            _, _, mort = all_results[sname]
            c = ws.cell(row=r, column=si, value=mort[comp])
            if comp == 'hysa_rate':
                c.number_format = PCT_FMT
            else:
                c.number_format = CUR_FMT
            c.alignment = Alignment(horizontal="center")
            c.border = THIN_BORDER
            if comp == 'total':
                c.font = HDR_FONT
                c.fill = PatternFill("solid", fgColor="E2EFDA")
        r += 1
    r += 1

    # ── Readiness gates ──
    r = _section_header(ws, r, "READINESS GATES (Base Scenario)")
    _, base_ms, base_mort = all_results.get("Base", list(all_results.values())[0])

    gates = [
        ("All debts paid off", base_ms.get('Fully Debt-Free')),
        ("$15,000 emergency fund preserved", base_ms.get('Emergency Fund Complete')),
        (f"HYSA reaches ${COMBINED_TARGET:,.0f} (e-fund + house)", base_ms.get('HYSA Reaches $75k')),
        ("Housing affordable vs. current rent",
         "YES" if base_mort['total'] <= CURRENT_RENT * 1.15 else "STRETCH"),
    ]
    for label, val in gates:
        ws.cell(row=r, column=1, value="✓" if val else "○").alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=label)
        c = ws.cell(row=r, column=3)
        if isinstance(val, date):
            c.value = val; c.number_format = "MMM YYYY"
            c.fill = MILE_FILL
        elif isinstance(val, str):
            c.value = val
            c.fill = MILE_FILL if val == "YES" else WARN_FILL
        else:
            c.value = "Not reached"; c.fill = WARN_FILL
        r += 1
    r += 1

    # ── Final readiness date ──
    r = _section_header(ws, r, "PROJECTED HOUSE-READY DATE")
    house_dt = base_ms.get('HYSA Reaches $75k')
    ws.cell(row=r, column=1, value="Earliest ready date:")
    c = ws.cell(row=r, column=2)
    if house_dt:
        c.value = house_dt; c.number_format = "MMMM YYYY"
        c.font = Font(name="Calibri", bold=True, size=14, color="006100")
        c.fill = MILE_FILL
    else:
        c.value = "Beyond forecast horizon"
        c.fill = WARN_FILL
    r += 2

    # ── Debt release waterfall ──
    r = _section_header(ws, r, "CASH-FLOW RELEASES AT DEBT FREEDOM")
    releases = [
        ("401(k) loan payroll deduction", f"${K401_DEDUCTION:,.2f}/paycheck",
         f"~${K401_DEDUCTION * 26 / 12:,.0f}/month"),
        ("Car loan payment", f"${CAR_PMT:,.2f}/month", ""),
        ("Student loan payments", f"${STU_PMT:,.2f}/month", ""),
        ("Bailey post-car contribution", f"${BAILEY_POST_CAR_SAVE:,.2f}/month",
         "starts after her car payoff"),
        ("TOTAL monthly release", "",
         f"~${K401_DEDUCTION * 26 / 12 + CAR_PMT + STU_PMT + BAILEY_POST_CAR_SAVE:,.0f}/month"),
    ]
    for label, amt, note in releases:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=amt)
        ws.cell(row=r, column=3, value=note)
        if label.startswith("TOTAL"):
            ws.cell(row=r, column=1).font = HDR_FONT
            ws.cell(row=r, column=2).font = HDR_FONT
            ws.cell(row=r, column=3).font = HDR_FONT
        r += 1
    r += 2

    # ── Source traceability ──
    r = _section_header(ws, r, "DATA SOURCES")
    sources = [
        "Debt balances & payoff sequence: finance/planning/goals.md (2026-06-23)",
        "Paycheck structure & saving power: finance/planning/cash-flow.md (2026-06-23)",
        "Pay cadence & profile: finance/profile.md",
        "Milestone history: finance/planning/actions-taken.md",
    ]
    for s in sources:
        ws.cell(row=r, column=1, value=s).font = Font(
            name="Calibri", size=9, italic=True, color="808080")
        r += 1

    return ws


def main():
    wb = Workbook()

    # Sheet 1: Assumptions
    build_assumptions_sheet(wb)

    # Run all three scenarios
    all_results = {}
    for scenario_name in SCENARIOS:
        rows, milestones, mortgage_info = run_forecast(scenario_name)
        all_results[scenario_name] = (rows, milestones, mortgage_info)

        # Sheet 2+: Forecast per scenario
        build_forecast_sheet(wb, rows, scenario_name)

    # Sheet: Mortgage (uses Base scenario)
    _, _, base_mort = all_results["Base"]
    build_mortgage_sheet(wb, base_mort)

    # Sheet: Summary
    build_summary_sheet(wb, all_results)

    # Save
    wb.save(OUTPUT)
    print(f"Workbook saved to: {OUTPUT}")

    # Print quick summary
    for name in SCENARIOS:
        _, ms, mort = all_results[name]
        print(f"\n── {name} Scenario ──")
        for k, v in ms.items():
            print(f"  {k}: {v.strftime('%b %Y') if v else 'Not reached'}")
        print(f"  Monthly housing: ${mort['total']:,.2f}")


if __name__ == "__main__":
    main()
